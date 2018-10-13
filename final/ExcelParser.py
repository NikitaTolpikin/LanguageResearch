import openpyxl
import json


openfile = 'project_data_filtered_united.json'

wb = openpyxl.load_workbook(filename='data.xlsx')
sheet = wb['test']

fh = open(openfile, 'r', encoding='utf-8')
project_data = json.load(fh)
fh.close()

i=2
j=0
sheet.cell(row=1, column=1).value = 'Город'
sheet.cell(row=1, column=2).value = 'Количество ошибок'
sheet.cell(row=1, column=3).value = 'Количество твитов'
sheet.cell(row=1, column=4).value = 'Количество слов'
sheet.cell(row=1, column=5).value = 'Количество неценз. выражений'
sheet.cell(row=1, column=6).value = 'Количество слов в твите'
sheet.cell(row=1, column=7).value = 'Ошибок на 1000 слов'
sheet.cell(row=1, column=8).value = 'Неценз. выражений на 1000 слов'
for city in project_data:
    sheet.cell(row=i, column=1).value = city
    sheet.cell(row=i, column=3).value = project_data[city]['twit count']
    sheet.cell(row=i, column=4).value = project_data[city]['word count']
    sheet.cell(row=i, column=5).value = project_data[city]['swear count']
    j=i+1
    for mistake in project_data[city]['mistakes']:
        sheet.cell(row=j, column=1).value = mistake
        sheet.cell(row=j, column=2).value = project_data[city]['mistakes'][mistake]
        j+=1
    sheet.cell(row=i, column=2).value = '=SUM(B' + str(i+1) + ':B' + str(j) + ')'
    sheet.cell(row=i, column=6).value = '=D' + str(i) + '/C' + str(i)
    sheet.cell(row=i, column=7).value = '=B' + str(i) + '/D' + str(i) + '*1000'
    sheet.cell(row=i, column=8).value = '=E' + str(i) + '/D' + str(i) + '*1000'
    i=j+1

wb.save('data.xlsx')
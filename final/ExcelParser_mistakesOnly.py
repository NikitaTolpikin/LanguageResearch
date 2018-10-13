import openpyxl
import json

openfile = 'project_data_filtered_united.json'

wb = openpyxl.load_workbook(filename='data_mistakes.xlsx')
sheet = wb['test']

fh = open(openfile, 'r', encoding='utf-8')
project_data = json.load(fh)
fh.close()
all_mistakes = {}

for city in project_data:
    for mistake in project_data[city]['mistakes']:
        if mistake in all_mistakes:
            all_mistakes[mistake] += project_data[city]['mistakes'][mistake]
        else:
            all_mistakes[mistake] = project_data[city]['mistakes'][mistake]

i=1
for mistake in all_mistakes:
    sheet.cell(row=i, column=1).value = mistake
    sheet.cell(row=i, column=2).value = all_mistakes[mistake]
    i+=1

wb.save('data_mistakes.xlsx')